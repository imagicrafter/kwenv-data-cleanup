#!/usr/bin/env python3
"""Check how many SF export customers (DKW + clean) are missing from geodata.js."""
import json, os, re, sys

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SF_EXPORT = '/Users/justinmartin/Downloads/sf-customer-list-full.xlsx'


def norm(name):
    stripped = re.sub(r'^DKW[\s\-\u2013]*', '', name, flags=re.I)
    return re.sub(r'[^A-Z0-9]', '', stripped.upper())


def main():
    # Load geodata
    geodata_path = os.path.join(SCRIPT_DIR, 'geodata.js')
    with open(geodata_path) as f:
        js = f.read()
    json_str = js.replace('const GEODATA = ', '').rstrip().rstrip(';')
    geodata = json.loads(json_str)
    geo_names = set(geodata['locations'].keys())
    print(f'Geodata locations: {len(geo_names)}')

    if not os.path.exists(SF_EXPORT):
        print(f'SF export not found at {SF_EXPORT}')
        sys.exit(1)

    wb = openpyxl.load_workbook(SF_EXPORT, read_only=True)
    ws = wb.active

    # Find header
    header_row = None
    headers = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        if row and str(row[0] or '').strip() == 'Customer Name':
            header_row = i + 1
            headers = [str(c or '').strip().lower() for c in row]
            break

    if not header_row:
        print('ERROR: Could not find header row')
        sys.exit(1)

    def find_col(keywords):
        for j, h in enumerate(headers):
            if all(k in h for k in keywords):
                return j
        return None

    col_addr = find_col(['primary', 'address 1']) or find_col(['address 1']) or find_col(['address'])

    dkw_total = 0
    dkw_in_geo = 0
    dkw_missing = 0
    clean_total = 0
    clean_in_geo = 0
    clean_missing = 0
    clean_missing_with_addr = 0
    clean_missing_names = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = str(row[0] or '').strip()
        if not name:
            continue

        n = norm(name)
        is_dkw = bool(re.match(r'^DKW[\s\-\u2013]', name, re.I))
        addr = str(row[col_addr] or '').strip() if col_addr is not None else ''

        if is_dkw:
            dkw_total += 1
            if n in geo_names:
                dkw_in_geo += 1
            else:
                dkw_missing += 1
        else:
            clean_total += 1
            if n in geo_names:
                clean_in_geo += 1
            else:
                clean_missing += 1
                if addr and len(addr) > 5:
                    clean_missing_with_addr += 1
                if len(clean_missing_names) < 25:
                    clean_missing_names.append((name, addr))

    wb.close()

    print(f'\n--- DKW Records ---')
    print(f'Total: {dkw_total}')
    print(f'In geodata: {dkw_in_geo}')
    print(f'Missing: {dkw_missing}')

    print(f'\n--- Clean (non-DKW) Records ---')
    print(f'Total: {clean_total}')
    print(f'In geodata: {clean_in_geo}')
    print(f'Missing: {clean_missing}')
    print(f'Missing with geocodable address: {clean_missing_with_addr}')

    if clean_missing_names:
        print(f'\nFirst {len(clean_missing_names)} missing clean customers:')
        for name, addr in clean_missing_names:
            print(f'  {name:40s} | {addr}')


if __name__ == '__main__':
    main()
