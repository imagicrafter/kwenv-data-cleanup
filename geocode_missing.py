#!/usr/bin/env python3
"""Geocode all SF records (DKW + clean) missing from geodata.js using Geocodio, then merge with territory assignment."""
import json, os, re, sys, time

try:
    import openpyxl
    import requests
except ImportError:
    print("ERROR: pip install openpyxl requests", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SF_EXPORT = None  # Set at runtime via CLI argument or SF_EXPORT_PATH env var
GEOCODIO_URL = 'https://api.geocod.io/v1.7/geocode'
BATCH_SIZE = 100  # Geocodio batch limit


def norm(name):
    stripped = re.sub(r'^DKW[\s\-\u2013]*', '', name, flags=re.I)
    return re.sub(r'[^A-Z0-9]', '', stripped.upper())


def load_env():
    d = SCRIPT_DIR
    for _ in range(10):
        env_file = os.path.join(d, '.env')
        if os.path.exists(env_file):
            with open(env_file) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        k, v = line.split('=', 1)
                        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
            return
        d = os.path.dirname(d)


def load_geodata():
    geodata_path = os.path.join(SCRIPT_DIR, 'geodata.js')
    with open(geodata_path) as f:
        js = f.read()
    json_str = js.replace('const GEODATA = ', '').rstrip().rstrip(';')
    return json.loads(json_str)


def parse_kml(path):
    from xml.etree import ElementTree as ET
    tree = ET.parse(path)
    root = tree.getroot()
    ns = 'http://www.opengis.net/kml/2.2'
    territories = []
    for pm in root.iter(f'{{{ns}}}Placemark'):
        name_el = pm.find(f'{{{ns}}}name')
        coords_el = pm.find(f'.//{{{ns}}}coordinates')
        if name_el is None or coords_el is None:
            continue
        name = name_el.text.strip()
        polygon = []
        for coord in coords_el.text.strip().split():
            parts = coord.split(',')
            lng, lat = float(parts[0]), float(parts[1])
            polygon.append((lng, lat))
        territories.append({'name': name, 'polygon': polygon})
    return territories


def point_in_polygon(lat, lng, polygon):
    n = len(polygon)
    inside = False
    j = n - 1
    for i in range(n):
        xi, yi = polygon[i]
        xj, yj = polygon[j]
        if ((yi > lat) != (yj > lat)) and (lng < (xj - xi) * (lat - yi) / (yj - yi) + xi):
            inside = not inside
        j = i
    return inside


def haversine(lat1, lng1, lat2, lng2):
    """Distance in km between two lat/lng points."""
    import math
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlng / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def nearest_territory(lat, lng, territories):
    """Find nearest territory by distance to polygon centroid. Returns (name, distance_km)."""
    best_name = None
    best_dist = float('inf')
    for t in territories:
        clat = sum(p[1] for p in t['polygon']) / len(t['polygon'])
        clng = sum(p[0] for p in t['polygon']) / len(t['polygon'])
        d = haversine(lat, lng, clat, clng)
        if d < best_dist:
            best_dist = d
            best_name = t['name']
    return best_name, best_dist


def get_missing_addresses(geodata):
    """Find ALL records (DKW + clean) in SF export that are NOT in geodata."""
    wb = openpyxl.load_workbook(SF_EXPORT, read_only=True)
    ws = wb.active

    geo_names = set(geodata['locations'].keys())
    missing = []
    seen_norms = set()

    # Find header row dynamically
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        if row and str(row[0] or '').strip() == 'Customer Name':
            header_row = i + 1
            headers = [str(c or '').strip().lower() for c in row]
            break

    if not header_row:
        print('ERROR: Could not find header row in SF export', file=sys.stderr)
        sys.exit(1)

    # Find address column indices
    def find_col(keywords):
        for j, h in enumerate(headers):
            if all(k in h for k in keywords):
                return j
        return None

    col_name = find_col(['customer', 'name'])
    col_addr = find_col(['primary', 'address 1']) or find_col(['address 1']) or find_col(['address'])
    col_city = find_col(['primary', 'city']) or find_col(['city'])
    col_state = find_col(['primary', 'state']) or find_col(['state'])
    col_zip = find_col(['primary', 'zip']) or find_col(['zip'])

    print(f'  Columns: name={col_name}, addr={col_addr}, city={col_city}, state={col_state}, zip={col_zip}')

    dkw_count = 0
    clean_count = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = str(row[col_name] or '').strip() if col_name is not None else ''
        if not name:
            continue
        n = norm(name)
        if n in geo_names or n in seen_norms:
            continue

        address = str(row[col_addr] or '').strip() if col_addr is not None else ''
        city = str(row[col_city] or '').strip() if col_city is not None else ''
        state = str(row[col_state] or '').strip() if col_state is not None else ''
        zipcode = str(row[col_zip] or '').strip() if col_zip is not None else ''

        if not address:
            continue

        full_address = address
        if city:
            full_address += ', ' + city
        if state:
            full_address += ', ' + state
        if zipcode:
            full_address += ' ' + zipcode

        is_dkw = bool(re.match(r'^DKW[\s\-\u2013]', name, re.I))
        # For display name, strip DKW prefix if present
        display_name = re.sub(r'^DKW[\s\-\u2013]*', '', name, flags=re.I) if is_dkw else name

        missing.append({
            'norm': n,
            'name': name,
            'display_name': display_name,
            'address': full_address,
        })
        seen_norms.add(n)

        if is_dkw:
            dkw_count += 1
        else:
            clean_count += 1

    wb.close()
    print(f'  Missing: {dkw_count} DKW + {clean_count} clean = {dkw_count + clean_count} total')
    return missing


CENSUS_URL = 'https://geocoding.geo.census.gov/geocoder/locations/onelineaddress'


def geocode_batch(addresses, api_key):
    """Geocode addresses using Geocodio, falling back to Census Bureau on quota limit."""
    results = {}
    failed = 0
    quota_hit = False

    for i, item in enumerate(addresses):
        if (i + 1) % 50 == 0 or i == 0:
            print(f'  [Geocodio] {i + 1}/{len(addresses)}...')
        try:
            resp = requests.get(
                GEOCODIO_URL,
                params={'q': item['address'], 'api_key': api_key},
                timeout=15,
            )
            if resp.status_code == 403:
                print(f'  Geocodio quota hit at record {i + 1}. Switching to Census Bureau...')
                quota_hit = True
                break
            if resp.status_code == 200:
                data = resp.json()
                if data.get('results') and len(data['results']) > 0:
                    loc = data['results'][0]['location']
                    results[item['norm']] = {
                        'lat': loc['lat'],
                        'lng': loc['lng'],
                        'accuracy': data['results'][0].get('accuracy', 0),
                    }
            else:
                failed += 1
        except Exception as e:
            failed += 1
            if failed <= 5:
                print(f'  Failed: {item["display_name"]}: {e}')
        if (i + 1) % 100 == 0:
            time.sleep(1)

    if quota_hit:
        remaining = [a for a in addresses[i:] if a['norm'] not in results]
        if remaining:
            print(f'  Geocoding {len(remaining)} remaining via Census Bureau (no daily limit)...')
            census_results = geocode_census_batch(remaining)
            results.update(census_results)

    return results


def geocode_census_batch(addresses):
    """Geocode addresses using the US Census Bureau geocoder (free, unlimited)."""
    results = {}
    failed = 0

    for i, item in enumerate(addresses):
        if (i + 1) % 50 == 0 or i == 0:
            print(f'  [Census] {i + 1}/{len(addresses)}...')
        try:
            resp = requests.get(
                CENSUS_URL,
                params={
                    'address': item['address'],
                    'benchmark': 'Public_AR_Current',
                    'format': 'json',
                },
                timeout=20,
            )
            if resp.status_code == 200:
                data = resp.json()
                matches = data.get('result', {}).get('addressMatches', [])
                if matches:
                    coords = matches[0]['coordinates']
                    results[item['norm']] = {
                        'lat': coords['y'],
                        'lng': coords['x'],
                        'accuracy': 0.8,
                    }
                else:
                    failed += 1
            else:
                failed += 1
        except Exception as e:
            failed += 1
            if failed <= 5:
                print(f'  Census failed: {item["display_name"]}: {e}')
        # Census rate limit is generous but add small delay to be respectful
        if (i + 1) % 50 == 0:
            time.sleep(0.5)

    print(f'  Census results: {len(results)} geocoded, {failed} failed')
    return results


def load_geocode_cache():
    """Load cached geocode results from previous runs."""
    cache_path = os.path.join(SCRIPT_DIR, 'geocode_cache.json')
    if os.path.exists(cache_path):
        with open(cache_path) as f:
            return json.load(f)
    return {}


def save_geocode_cache(cache):
    """Save geocode cache to disk."""
    cache_path = os.path.join(SCRIPT_DIR, 'geocode_cache.json')
    with open(cache_path, 'w') as f:
        json.dump(cache, f, separators=(',', ':'))
    print(f'  Cache saved: {len(cache)} results → geocode_cache.json')


def save_geodata(geodata):
    """Write geodata.js to disk."""
    output_path = os.path.join(SCRIPT_DIR, 'geodata.js')
    with open(output_path, 'w') as f:
        f.write('const GEODATA = ')
        json.dump(geodata, f, separators=(',', ':'))
        f.write(';\n')
    size_kb = os.path.getsize(output_path) / 1024
    print(f'  Wrote geodata.js ({size_kb:.1f} KB, {len(geodata["locations"])} locations)')


def assign_territory(lat, lng, territories):
    """Assign a territory via point-in-polygon, falling back to nearest."""
    for t in territories:
        if point_in_polygon(lat, lng, t['polygon']):
            return t['name'], 'polygon'
    if territories:
        name, _ = nearest_territory(lat, lng, territories)
        return name, 'nearest'
    return None, None


def main():
    import argparse
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('sf_export', nargs='?', help='Path to SF customer list export .xlsx (overrides SF_EXPORT_PATH env var)')
    args = parser.parse_args()

    load_env()

    sf_export = args.sf_export or os.environ.get('SF_EXPORT_PATH')
    if not sf_export:
        print('ERROR: Provide the SF export path as an argument or set SF_EXPORT_PATH env var.', file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(sf_export):
        print(f'ERROR: SF export not found at {sf_export}. Set SF_EXPORT_PATH env var or pass as argument.', file=sys.stderr)
        sys.exit(1)

    # Patch module-level name used by get_missing_addresses
    global SF_EXPORT
    SF_EXPORT = sf_export

    api_key = os.environ.get('GEOCODIO_API_KEY')
    if not api_key:
        print('ERROR: GEOCODIO_API_KEY required in .env', file=sys.stderr)
        sys.exit(1)

    # Load existing geodata
    geodata = load_geodata()
    print(f'Existing geodata: {len(geodata["locations"])} locations')

    # Load geocode cache from previous runs
    cache = load_geocode_cache()
    if cache:
        print(f'Geocode cache: {len(cache)} cached results from previous runs')

    # Load KML territories for point-in-polygon
    kml_path = os.path.join(SCRIPT_DIR, 'territories.kml')
    territories = parse_kml(kml_path) if os.path.exists(kml_path) else []
    print(f'Territory polygons: {len(territories)}')

    # Apply any cached geocodes that aren't yet in geodata
    cache_applied = 0
    for norm_key, geo in cache.items():
        if norm_key not in geodata['locations']:
            entry = {'n': geo.get('display_name', ''), 'lat': geo['lat'], 'lng': geo['lng']}
            terr, method = assign_territory(geo['lat'], geo['lng'], territories)
            if terr:
                entry['t'] = terr
                if method == 'nearest':
                    entry['a'] = 'nearest'
            geodata['locations'][norm_key] = entry
            cache_applied += 1
    if cache_applied:
        print(f'Applied {cache_applied} cached geocodes to geodata')
        save_geodata(geodata)

    # Find missing addresses
    missing = get_missing_addresses(geodata)
    print(f'Missing records needing geocoding: {len(missing)}')

    if not missing:
        print('Nothing to geocode!')
        return

    # Filter out records with no useful address
    geocodable = [m for m in missing if len(m['address']) > 10]
    skipped = len(missing) - len(geocodable)
    if skipped:
        print(f'Skipping {skipped} records with insufficient address data')
    print(f'Geocoding {len(geocodable)} addresses via Geocodio...')

    # Geocode in chunks of 100, saving after each chunk
    CHUNK = 100
    total_geocoded = 0
    polygon_assigned = 0
    nearest_assigned = 0
    quota_hit = False

    for start in range(0, len(geocodable), CHUNK):
        chunk = geocodable[start:start + CHUNK]
        geocoded = geocode_batch(chunk, api_key)
        total_geocoded += len(geocoded)

        # Update cache and geodata for this chunk
        for item in chunk:
            geo = geocoded.get(item['norm'])
            if not geo:
                continue

            # Save to cache (raw lat/lng + display name)
            cache[item['norm']] = {
                'lat': geo['lat'],
                'lng': geo['lng'],
                'display_name': item['display_name'],
            }

            # Build geodata entry with territory assignment
            entry = {'n': item['display_name'], 'lat': geo['lat'], 'lng': geo['lng']}
            terr, method = assign_territory(geo['lat'], geo['lng'], territories)
            if terr:
                entry['t'] = terr
                if method == 'nearest':
                    entry['a'] = 'nearest'
                    nearest_assigned += 1
                else:
                    polygon_assigned += 1

            if item['norm'] not in geodata['locations']:
                geodata['locations'][item['norm']] = entry

        # Save after every chunk
        save_geocode_cache(cache)
        save_geodata(geodata)

        # If we got fewer results than the chunk size, we likely hit quota
        if len(geocoded) < len(chunk):
            quota_hit = True
            break

    # If Geocodio quota hit, fall back to Census for all remaining
    if quota_hit:
        remaining = [a for a in geocodable if a['norm'] not in cache]
        if remaining:
            print(f'\nGeocodio quota exhausted. Falling back to Census Bureau for {len(remaining)} remaining...')
            census_results = geocode_census_batch(remaining)
            total_geocoded += len(census_results)

            for item in remaining:
                geo = census_results.get(item['norm'])
                if not geo:
                    continue
                cache[item['norm']] = {
                    'lat': geo['lat'],
                    'lng': geo['lng'],
                    'display_name': item['display_name'],
                }
                entry = {'n': item['display_name'], 'lat': geo['lat'], 'lng': geo['lng']}
                terr, method = assign_territory(geo['lat'], geo['lng'], territories)
                if terr:
                    entry['t'] = terr
                    if method == 'nearest':
                        entry['a'] = 'nearest'
                        nearest_assigned += 1
                    else:
                        polygon_assigned += 1
                if item['norm'] not in geodata['locations']:
                    geodata['locations'][item['norm']] = entry

            save_geocode_cache(cache)
            save_geodata(geodata)

    print(f'\nGeocoded this run: {total_geocoded}')
    print(f'  Polygon-assigned: {polygon_assigned}')
    print(f'  Nearest-assigned: {nearest_assigned}')
    print(f'Total locations now: {len(geodata["locations"])}')
    print(f'Total cached geocodes: {len(cache)}')


if __name__ == '__main__':
    main()
